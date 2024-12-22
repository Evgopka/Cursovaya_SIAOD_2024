import datetime
import random
import openpyxl

# дни недели
weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
# выходные дни
weekend = ["Saturday", "Sunday"]

# рабочие часы для водителя 8h
work_8h = 8
# рабочие часы для водителя 12h
work_12h = 12
# время обеда в минутах для водителя 8h
lunch_8h = 60
# мин время перерыва в минутах для водителя 12h
break_12h = 15
# частота перерывов для водителя 12h в минутах
break_freq_12h = 120
# время длинного перерыва для водителя 12h
long_break_12h = 40
# мин время на маршрут в минутах
route_min = 50
# макс время на маршрут в минутах
route_max = 70
# мин время на пересмену в минутах
change_min = 10
# макс время на пересмену в минутах
change_max = 15
# общий поток пассажиров
pass_flow = 1000
# процент пассажиров в час пик
peak_pass_percent = 0.7
# начало смены
start_time = datetime.datetime(2005, 1, 31, 6, 0) #Это мой день рождения ;)
# конец смены
end_time = datetime.datetime(2005, 2, 1, 3, 0)
# начало первого часа пик
peak_start_1 = datetime.datetime(2005, 1, 31, 7, 0)
# конец первого часа пик
peak_end_1 = datetime.datetime(2005, 1, 31, 9, 0)
# начало второго часа пик
peak_start_2 = datetime.datetime(2005, 1, 31, 17, 0)
# конец второго часа пик
peak_end_2 = datetime.datetime(2005, 1, 31, 19, 0)

# размер популяции для генетического алгоритма
pop_size = 50
# количество поколений для генетического алгоритма
gens = 100
# вероятность мутации для генетического алгоритма
mut_rate = 0.1




class person:
    def __init__(self, type, id):
        self.type = type # тип водителя
        self.schedule = [] # расписание водителя
        self.total_work_time = datetime.timedelta() # общее время работы водителя
        self.last_break = start_time # время последнего перерыва водителя
        self.id = id # id водителя
        self.lunch_taken = False # флаг взят ли обед

    def __repr__(self):
        return f"person(id={self.id}, type={self.type}, schedule={len(self.schedule)} shifts, worktime = {self.total_work_time})"


class way:
    def __init__(self, start_time, route_time, driver_id):
        self.start_time = start_time # время начала маршрута
        self.end_time = start_time + datetime.timedelta(minutes=route_time) # время конца маршрута
        self.driver_id = driver_id # id водителя назначенного на маршрут

    def __repr__(self):
        return f"way(start_time={self.start_time.strftime('%H:%M')}, end_time={self.end_time.strftime('%H:%M')}, driver_id={self.driver_id})"


class group:
    def __init__(self):
        self.routes = [] # список маршрутов
        self.drivers = [] # список водителей

    def add_route(self, route):
        self.routes.append(route)

    def add_driver(self, driver):
        self.drivers.append(driver)

    def calculate_metrics(self):
        peak_routes = 0
        for route in self.routes:
            if (
                    (route.start_time >= peak_start_1 and route.start_time < peak_end_1) or
                    (route.start_time >= peak_start_2 and route.start_time < peak_end_2)
            ):
                peak_routes += 1
        unique_drivers = len(self.drivers)
        total_routes = len(self.routes)
        return total_routes, peak_routes, unique_drivers


def is_peak(time):
    return (time >= peak_start_1 and time < peak_end_1) or (
            time >= peak_start_2 and time < peak_end_2)


def is_weekend(date):
    return date.strftime('%A') in weekend

def create_straight(num_buses, num_drivers_a, num_drivers_b, current_date):
    schedule = group()
    drivers_a = []
    drivers_b = []
    current_time = start_time

    for i in range(num_drivers_a):
        drivers_a.append(person('8h', f'8h_{i + 1}'))
    for i in range(num_drivers_b):
        drivers_b.append(person('12h', f'12h_{i + 1}'))

    available_drivers_a = list(drivers_a)
    available_drivers_b = list(drivers_b)

    while current_time < end_time:
        route_time = random.randint(route_min, route_max)
        if is_peak(current_time) and not is_weekend(current_date):

            for _ in range(int(num_buses * peak_pass_percent)):
                if available_drivers_a:
                    driver = available_drivers_a[0]
                    if driver.type == '8h' and current_time >= start_time + datetime.timedelta(
                            hours=4) and not driver.lunch_taken:
                        lunch_end_time = current_time + datetime.timedelta(minutes=lunch_8h)
                        driver.schedule.append((current_time, lunch_end_time, 'lunch'))
                        driver.total_work_time += datetime.timedelta(minutes=lunch_8h)
                        driver.lunch_taken = True
                        current_time = lunch_end_time
                        continue
                    if driver.total_work_time + datetime.timedelta(minutes=route_time) <= datetime.timedelta(
                            hours=work_8h):
                        end_time_route = current_time + datetime.timedelta(minutes=route_time)
                        if end_time_route <= end_time:
                            route = way(current_time, route_time, driver.id)
                            schedule.add_route(route)
                            driver.schedule.append((route.start_time, route.end_time, 'route'))
                            driver.total_work_time += datetime.timedelta(minutes=route_time)
                            current_time = end_time_route + datetime.timedelta(
                                minutes=random.randint(change_min,
                                                       change_max))
                        else:
                            available_drivers_a.pop(0)
                            continue
                    else:
                        available_drivers_a.pop(0)
                        continue
                elif available_drivers_b:
                    driver = available_drivers_b[0]
                    if driver.total_work_time >= datetime.timedelta(
                            minutes=break_freq_12h) and driver.last_break <= current_time - datetime.timedelta(
                        minutes=break_freq_12h):
                        break_start_time = current_time
                        break_end_time = current_time + datetime.timedelta(minutes=long_break_12h)
                        driver.schedule.append((break_start_time, break_end_time, 'break'))
                        driver.total_work_time += datetime.timedelta(minutes=long_break_12h)
                        driver.last_break = break_end_time
                        current_time += datetime.timedelta(minutes=long_break_12h)
                        continue
                    else:
                        end_time_route = current_time + datetime.timedelta(minutes=route_time)
                        if end_time_route <= end_time:
                            route = way(current_time, route_time, driver.id)
                            schedule.add_route(route)
                            driver.schedule.append((route.start_time, route.end_time, 'route'))
                            driver.total_work_time += datetime.timedelta(minutes=route_time)
                            current_time = end_time_route + datetime.timedelta(
                                minutes=random.randint(change_min,
                                                       change_max))
                        else:
                            continue
                else:
                    break
        else:
            passenger_percent = 1 - peak_pass_percent if not is_weekend(current_date) else 1
            for _ in range(int(num_buses * passenger_percent)):
                if available_drivers_a:
                    driver = available_drivers_a[0]
                    if driver.type == '8h' and current_time >= start_time + datetime.timedelta(
                            hours=4) and not driver.lunch_taken:
                        lunch_end_time = current_time + datetime.timedelta(minutes=lunch_8h)
                        driver.schedule.append((current_time, lunch_end_time, 'lunch'))
                        driver.total_work_time += datetime.timedelta(minutes=lunch_8h)
                        driver.lunch_taken = True
                        current_time = lunch_end_time
                        continue
                    if driver.total_work_time + datetime.timedelta(minutes=route_time) <= datetime.timedelta(
                            hours=work_8h):
                        end_time_route = current_time + datetime.timedelta(minutes=route_time)
                        if end_time_route <= end_time:
                            route = way(current_time, route_time, driver.id)
                            schedule.add_route(route)
                            driver.schedule.append((route.start_time, route.end_time, 'route'))
                            driver.total_work_time += datetime.timedelta(minutes=route_time)
                            current_time = end_time_route + datetime.timedelta(
                                minutes=random.randint(change_min,
                                                       change_max))
                        else:
                            available_drivers_a.pop(0)
                            continue
                    else:
                        available_drivers_a.pop(0)
                        continue
                elif available_drivers_b:
                    driver = available_drivers_b[0]
                    if driver.total_work_time >= datetime.timedelta(
                            minutes=break_freq_12h) and driver.last_break <= current_time - datetime.timedelta(
                        minutes=break_freq_12h):
                        break_start_time = current_time
                        break_end_time = current_time + datetime.timedelta(minutes=long_break_12h)
                        driver.schedule.append((break_start_time, break_end_time, 'break'))
                        driver.total_work_time += datetime.timedelta(minutes=long_break_12h)
                        driver.last_break = break_end_time
                        current_time += datetime.timedelta(minutes=long_break_12h)
                        continue
                    else:
                        end_time_route = current_time + datetime.timedelta(minutes=route_time)
                        if end_time_route <= end_time:
                            route = way(current_time, route_time, driver.id)
                            schedule.add_route(route)
                            driver.schedule.append((route.start_time, route.end_time, 'route'))
                            driver.total_work_time += datetime.timedelta(minutes=route_time)
                            current_time = end_time_route + datetime.timedelta(
                                minutes=random.randint(change_min,
                                                       change_max))
                        else:
                            continue
                else:
                    break
    schedule.drivers.extend(drivers_a)
    schedule.drivers.extend(drivers_b)
    return schedule

def generate_random(num_buses, num_drivers_a, num_drivers_b, current_date):
    schedule = group()
    drivers = []
    for i in range(num_drivers_a):
        drivers.append(person('8h', f'8h_{i + 1}'))
    for i in range(num_drivers_b):
        drivers.append(person('12h', f'12h_{i + 1}'))

    current_time = start_time

    while current_time < end_time:
        route_time = random.randint(route_min, route_max)
        if is_peak(current_time) and not is_weekend(current_date):
            for _ in range(int(num_buses * peak_pass_percent)):
                if drivers:
                    driver = random.choice(drivers)
                    if driver.type == '8h' and driver.total_work_time + datetime.timedelta(
                            minutes=route_time) <= datetime.timedelta(hours=work_8h):
                        if current_time >= start_time + datetime.timedelta(hours=4) and not driver.lunch_taken:
                            lunch_end_time = current_time + datetime.timedelta(minutes=lunch_8h)
                            driver.schedule.append((current_time, lunch_end_time, 'lunch'))
                            driver.total_work_time += datetime.timedelta(minutes=lunch_8h)
                            driver.lunch_taken = True
                            current_time = lunch_end_time
                            continue
                        else:
                            end_time_route = current_time + datetime.timedelta(minutes=route_time)
                            if end_time_route <= end_time:
                                route = way(current_time, route_time, driver.id)
                                schedule.add_route(route)
                                driver.schedule.append((route.start_time, route.end_time, 'route'))
                                driver.total_work_time += datetime.timedelta(minutes=route_time)
                                current_time = end_time_route + datetime.timedelta(
                                    minutes=random.randint(change_min, change_max))
                            else:
                                continue
                    elif driver.type == '12h':

                        if driver.total_work_time >= datetime.timedelta(
                                minutes=break_freq_12h) and driver.last_break <= current_time - datetime.timedelta(
                            minutes=break_freq_12h):
                            break_start_time = current_time
                            break_end_time = current_time + datetime.timedelta(minutes=long_break_12h)
                            driver.schedule.append((break_start_time, break_end_time, 'break'))
                            driver.total_work_time += datetime.timedelta(minutes=long_break_12h)
                            driver.last_break = break_end_time
                            current_time += datetime.timedelta(minutes=long_break_12h)
                            continue
                        else:
                            end_time_route = current_time + datetime.timedelta(minutes=route_time)
                            if end_time_route <= end_time:
                                route = way(current_time, route_time, driver.id)
                                schedule.add_route(route)
                                driver.schedule.append((route.start_time, route.end_time, 'route'))
                                driver.total_work_time += datetime.timedelta(minutes=route_time)
                                current_time = end_time_route + datetime.timedelta(
                                    minutes=random.randint(change_min, change_max))
                            else:
                                continue
                    else:
                        drivers.remove(driver)
                        continue
                else:
                    break

        else:
            passenger_percent = 1 - peak_pass_percent if not is_weekend(current_date) else 1
            for _ in range(int(num_buses * passenger_percent)):
                if drivers:
                    driver = random.choice(drivers)
                    if driver.type == '8h' and driver.total_work_time + datetime.timedelta(
                            minutes=route_time) <= datetime.timedelta(hours=work_8h):
                        if current_time >= start_time + datetime.timedelta(hours=4) and not driver.lunch_taken:
                            lunch_end_time = current_time + datetime.timedelta(minutes=lunch_8h)
                            driver.schedule.append((current_time, lunch_end_time, 'lunch'))
                            driver.total_work_time += datetime.timedelta(minutes=lunch_8h)
                            driver.lunch_taken = True
                            current_time = lunch_end_time
                            continue
                        else:
                            end_time_route = current_time + datetime.timedelta(minutes=route_time)
                            if end_time_route <= end_time:
                                route = way(current_time, route_time, driver.id)
                                schedule.add_route(route)
                                driver.schedule.append((route.start_time, route.end_time, 'route'))
                                driver.total_work_time += datetime.timedelta(minutes=route_time)
                                current_time = end_time_route + datetime.timedelta(
                                    minutes=random.randint(change_min, change_max))
                            else:
                                continue
                    elif driver.type == '12h':
                        if driver.total_work_time >= datetime.timedelta(minutes=break_freq_12h) and driver.last_break <= current_time - datetime.timedelta(minutes=break_freq_12h):
                            break_start_time = current_time
                            break_end_time = current_time + datetime.timedelta(minutes=long_break_12h)
                            driver.schedule.append((break_start_time, break_end_time, 'break'))
                            driver.total_work_time += datetime.timedelta(minutes=long_break_12h)
                            driver.last_break = break_end_time
                            current_time += datetime.timedelta(minutes=long_break_12h)
                            continue
                        else:
                            end_time_route = current_time + datetime.timedelta(minutes=route_time)
                            if end_time_route <= end_time:
                                route = way(current_time, route_time, driver.id)
                                schedule.add_route(route)
                                driver.schedule.append((route.start_time, route.end_time, 'route'))
                                driver.total_work_time += datetime.timedelta(minutes=route_time)
                                current_time = end_time_route + datetime.timedelta(
                                    minutes=random.randint(change_min, change_max))
                            else:
                                continue
                    else:
                        drivers.remove(driver)
                        continue
                else:
                    break
        current_time += datetime.timedelta(minutes=route_time + random.randint(change_min, change_max))

    schedule.drivers.extend(drivers)
    return schedule

def fitness(schedule):
    total_routes, peak_routes, unique_drivers = schedule.calculate_metrics()
    return total_routes - unique_drivers * 0.1


def crossover(schedule1, schedule2):
    split_point = random.randint(0, min(len(schedule1.routes), len(schedule2.routes)))
    child_schedule = group()
    child_schedule.routes = schedule1.routes[:split_point] + schedule2.routes[split_point:]

    split_point = random.randint(0, min(len(schedule1.drivers), len(schedule2.drivers)))
    child_schedule.drivers = schedule1.drivers[:split_point] + schedule2.drivers[split_point:]
    return child_schedule


def mutate(schedule):
    if random.random() < mut_rate:
        if schedule.routes:
            index_route_mutate = random.randint(0, len(schedule.routes) - 1)
            new_start_time = schedule.routes[index_route_mutate].start_time + datetime.timedelta(minutes=random.randint(-30, 30))
            if new_start_time > start_time and new_start_time < end_time:
                schedule.routes[index_route_mutate] = way(new_start_time, random.randint(route_min, route_max),
                                                          schedule.routes[index_route_mutate].driver_id)
        if schedule.drivers:
            index_driver_mutate = random.randint(0, len(schedule.drivers) - 1)
            schedule.drivers[index_driver_mutate].type = random.choice(['8h', '12h'])
    return schedule


def genetic_algorithm(num_buses, num_drivers_a, num_drivers_b, current_date):
    population = [generate_random(num_buses, num_drivers_a, num_drivers_b, current_date) for _ in range(pop_size)]

    for generation in range(gens):
        population.sort(key=fitness, reverse=True)
        parents = population[:pop_size // 2]

        offspring = []
        for i in range(0, len(parents), 2):
            if i + 1 < len(parents):
                child1 = crossover(parents[i], parents[i + 1])
                child2 = crossover(parents[i + 1], parents[i])
                offspring.append(mutate(child1))
                offspring.append(mutate(child2))
            else:
                offspring.append(mutate(parents[i]))

        population = parents + offspring
        population.sort(key=fitness, reverse=True)
        population = population[:pop_size]

    return population[0]


def write_schedule_to_excel(straight_schedules, genetic_schedules, filename, start_date):
    workbook = openpyxl.Workbook()

    days = [start_date + datetime.timedelta(days=i) for i in range(7)]
    day_names = ["Понедельник", "Вторник", "Среда", "Четверг", "пятница", "Суббота", "Воскресенье"]

    for day, day_name in zip(days, day_names):
        sheet = workbook.create_sheet(title=day_name)
        sheet.append(['Тип алгоритма', 'Имя водителя', 'Расписание', 'Перерывы'])

        straight_schedule = straight_schedules[day_name]
        genetic_schedule = genetic_schedules[day_name]

        for schedule, algorithm_name in [(straight_schedule, "В лоб"), (genetic_schedule, "Генетический")]:
            for driver in schedule.drivers:
                shifts_text = ""
                breaks_text = ""
                for start, end, type in driver.schedule:
                    start_datetime = datetime.datetime.combine(day, start.time())
                    end_datetime = datetime.datetime.combine(day, end.time())
                    if type == 'route':
                        shifts_text += f"{start_datetime.strftime('%H:%M')}-{end_datetime.strftime('%H:%M')}, "
                    elif type == 'break' or type == 'lunch':
                        breaks_text += f"{start_datetime.strftime('%H:%M')}-{end_datetime.strftime('%H:%M')}, "
                shifts_text = shifts_text.rstrip(", ")
                breaks_text = breaks_text.rstrip(", ")
                sheet.append([algorithm_name, driver.id, shifts_text, breaks_text])

    if "Sheet" in workbook.sheetnames:
        std = workbook["Sheet"]
        workbook.remove(std)
    workbook.save(filename)

def run_algorithms_and_save(num_buses, num_drivers_a, num_drivers_b, current_date):
    days = [current_date + datetime.timedelta(days=i) for i in range(7)]
    day_names = ["Понедельник", "Вторник", "Среда", "Четверг", "пятница", "Суббота", "Воскресенье"]

    straight_schedules = {}
    genetic_schedules = {}

    for day, day_name in zip(days, day_names):
        straight_schedules[day_name] = create_straight(num_buses, num_drivers_a, num_drivers_b, day)
        genetic_schedules[day_name] = genetic_algorithm(num_buses, num_drivers_a, num_drivers_b, day)

    filename = 'расписание_маршруток.xlsx'
    write_schedule_to_excel(straight_schedules, genetic_schedules, filename, current_date)

    print(f"Расписание сохранено в {filename}")


if __name__ == "__main__":
    num_buses = 8
    num_drivers_a = 10
    num_drivers_b = 5
    current_date = datetime.date(2024, 12, 22)

    run_algorithms_and_save(num_buses, num_drivers_a, num_drivers_b, current_date)